"""
QLA Excel parser and deep analysis engine.
"""

import re
import statistics
from datetime import datetime
from pathlib import Path

import openpyxl

GROUP_MAP = {
    "CP":        ("Core Plus",  "Middle", "short_answer"),
    "N":         ("Numeracy",   "Lower",  "multiple_choice"),
    "EXTENSION": ("Extension",  "Higher", "exam_style"),
    "EXT":       ("Extension",  "Higher", "exam_style"),
}
SCORE_TOPICS = {"paper 1 score", "paper 2 score", "total score", "grade"}


# ── Shared parsers ────────────────────────────────────────────────────────────

def detect_groups(sheet_names):
    found = {}
    for name in sheet_names:
        if "student grid" in name.lower():
            continue
        upper = name.upper()
        for code, info in GROUP_MAP.items():
            if f" {code} " in upper or f"- {code} " in upper or f" {code}-" in upper:
                found.setdefault(code, info)
    return found


def _find_anchor(all_rows):
    for row_idx, row in enumerate(all_rows[:15]):
        for col_idx, val in enumerate(row):
            if val == "Question number":
                return row_idx, col_idx + 1
    return None, None


def _sheets_for_group(wb, code):
    return [
        s for s in wb.sheetnames
        if "student grid" not in s.lower()
        and (f" {code} " in s.upper() or f"- {code} " in s.upper() or f" {code}-" in s.upper())
    ]


def parse_sheet(ws, class_filter=None):
    all_rows = list(ws.iter_rows(values_only=True))
    qnum_row_idx, q_start_col = _find_anchor(all_rows)
    if qnum_row_idx is None:
        return None

    topic_row  = all_rows[qnum_row_idx - 1]
    header_row = all_rows[qnum_row_idx + 1]
    data_start = qnum_row_idx + 2
    class_col  = q_start_col - 1

    questions = []
    for col in range(q_start_col, len(topic_row)):
        topic = topic_row[col] if col < len(topic_row) else None
        marks = header_row[col] if col < len(header_row) else None
        if topic is None and marks is None:
            break
        if not isinstance(marks, (int, float)):
            continue
        if isinstance(topic, str) and topic.strip().lower() in SCORE_TOPICS:
            break
        questions.append({
            "topic":        str(topic).strip() if topic else f"Q{col}",
            "max_marks":    float(marks),
            "col":          col,
            "scores":       [],
            "class_scores": {},
        })

    if not questions:
        return None

    for row in all_rows[data_start:]:
        cls = row[class_col] if class_col < len(row) else None
        if not isinstance(cls, str):
            continue
        if class_filter and cls not in class_filter:
            continue
        for q in questions:
            score = row[q["col"]] if q["col"] < len(row) else None
            if isinstance(score, (int, float)):
                q["scores"].append(score)
                q["class_scores"].setdefault(cls, []).append(score)

    return questions


def merge_questions(sheet_q_lists):
    merged = {}
    for sheet_qs in sheet_q_lists:
        for q in sheet_qs:
            key = q["topic"]
            if key not in merged:
                merged[key] = {
                    "topic":        q["topic"],
                    "max_marks":    q["max_marks"],
                    "scores":       [],
                    "class_scores": {},
                }
            merged[key]["scores"].extend(q["scores"])
            for cls, sc in q["class_scores"].items():
                merged[key]["class_scores"].setdefault(cls, []).extend(sc)
    return list(merged.values())


def get_all_classes(wb, sheet_names):
    classes = set()
    for name in sheet_names:
        ws = wb[name]
        all_rows = list(ws.iter_rows(values_only=True))
        qnum_row_idx, q_start_col = _find_anchor(all_rows)
        if qnum_row_idx is None:
            continue
        class_col  = q_start_col - 1
        data_start = qnum_row_idx + 2
        for row in all_rows[data_start:]:
            cls = row[class_col] if class_col < len(row) else None
            if isinstance(cls, str):
                classes.add(cls)
    return sorted(classes)


def _parse_student_grid(wb, sheet_name):
    """
    Read a Student Grids sheet and return a list of topics with max marks.
    These sheets hold the self-assessment structure even when score sheets are empty.
    """
    try:
        ws = wb[sheet_name]
    except KeyError:
        return []
    all_rows = list(ws.iter_rows(values_only=True))
    topics = []
    for row in all_rows[1:]:  # skip header row
        # Paper 1 columns: 0=Q, 1=topic, 3=out_of
        if len(row) > 3 and row[1] and isinstance(row[1], str) and isinstance(row[3], (int, float)):
            topics.append({"topic": str(row[1]).strip(), "max_marks": float(row[3]), "paper": 1})
        # Paper 2 columns: 8=Q, 9=topic, 11=out_of
        if len(row) > 11 and row[9] and isinstance(row[9], str) and isinstance(row[11], (int, float)):
            topics.append({"topic": str(row[9]).strip(), "max_marks": float(row[11]), "paper": 2})
    return topics


def _pct(scores, max_marks):
    if not scores or max_marks <= 0:
        return None
    return round(sum(scores) / (len(scores) * max_marks) * 100, 1)


# ── Stats helper ──────────────────────────────────────────────────────────────

def _topic_stats(scores, max_marks):
    """Return detailed statistics dict for a set of raw scores on one topic."""
    if not scores or max_marks <= 0:
        return None
    n = len(scores)
    avg_raw = sum(scores) / n
    avg_pct = round(avg_raw / max_marks * 100, 1)
    std_dev = round(statistics.stdev(scores), 2) if n > 1 else 0.0
    pct_full = round(sum(1 for s in scores if s >= max_marks) / n * 100, 1)
    pct_zero = round(sum(1 for s in scores if s == 0) / n * 100, 1)
    return {
        "avg_pct":      avg_pct,
        "avg_raw":      round(avg_raw, 2),
        "std_dev":      std_dev,
        "pct_full":     pct_full,
        "pct_zero":     pct_zero,
        "num_students": n,
    }


def _score_band(pct):
    """Return a CSS band class for a percentage score."""
    if pct is None:
        return "band-none"
    if pct < 30:
        return "band-vlow"
    if pct < 50:
        return "band-low"
    if pct < 65:
        return "band-mid"
    if pct < 80:
        return "band-good"
    return "band-high"


# ── Public API ────────────────────────────────────────────────────────────────

def scan_qla(filepath):
    """Lightweight scan: groups, sheets, and classes detected in the file."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    detected = detect_groups(wb.sheetnames)
    m = re.search(r"Year[\s_]*(\d+)", str(filepath), re.I)
    year_group = m.group(1) if m else "?"

    result = {"year_group": year_group, "groups": {}}
    for code, info in detected.items():
        sheets  = _sheets_for_group(wb, code)
        classes = get_all_classes(wb, sheets)
        result["groups"][code] = {
            "name": info[0], "ability": info[1], "format": info[2],
            "sheets": sheets, "classes": classes,
        }
    return result


def deep_analyse_qla(filepath, question_bank=None):
    """
    Full deep analysis of a QLA file.

    Returns a rich dict containing:
      - per-group, per-class topic stats (avg%, std_dev, %full, %zero)
      - cross-class heatmap data per group
      - class rankings within each group
      - cross-group universally weak topics
      - cohort-level summary stats
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    detected = detect_groups(wb.sheetnames)
    m = re.search(r"Year[\s_]*(\d+)", str(filepath), re.I)
    year_group = m.group(1) if m else "?"

    output = {
        "filename":     Path(filepath).name,
        "year_group":   year_group,
        "analysed_at":  datetime.now().strftime("%d %B %Y, %H:%M"),
        "groups":       {},
    }

    for code, (group_name, ability, fmt) in detected.items():
        sheets      = _sheets_for_group(wb, code)
        all_classes = get_all_classes(wb, sheets)

        # Merge all sheets without a class filter to get full topic/class_scores data
        all_sheet_qs = [qs for s in sheets for qs in [parse_sheet(wb[s])] if qs]
        merged       = merge_questions(all_sheet_qs)

        # Detect if this file has any student score data at all
        has_scores = any(q["scores"] for q in merged)

        # ── Fallback: read topic list from Student Grids if no scores ────
        if not has_scores:
            grid_sheet = next((s for s in wb.sheetnames if "student grid" in s.lower() and code in s.upper()), None)
            grid_topics = _parse_student_grid(wb, grid_sheet) if grid_sheet else []
        else:
            grid_topics = []

        # ── Per-class detailed stats (only when scores exist) ────────────
        classes_data = {}
        for cls in all_classes:
            cls_topic_stats = []
            for q in merged:
                scores = q["class_scores"].get(cls, [])
                stats  = _topic_stats(scores, q["max_marks"])
                if stats:
                    cls_topic_stats.append({
                        "topic":         q["topic"],
                        "max_marks":     q["max_marks"],
                        **stats,
                        "band":          _score_band(stats["avg_pct"]),
                        "question_text": (question_bank or {}).get(q["topic"], {}).get("text"),
                    })

            cls_topic_stats.sort(key=lambda x: x["avg_pct"])
            avgs = [t["avg_pct"] for t in cls_topic_stats]
            overall = round(sum(avgs) / len(avgs), 1) if avgs else None
            student_count = max((t["num_students"] for t in cls_topic_stats), default=0)

            classes_data[cls] = {
                "student_count":   student_count,
                "overall_avg_pct": overall,
                "topics":          cls_topic_stats,
                "weakest_3":       cls_topic_stats[:3],
                "strongest_3":     list(reversed(cls_topic_stats[-3:])),
                "band":            _score_band(overall),
            }

        # ── Group average and class rankings ─────────────────────────────
        avgs = [d["overall_avg_pct"] for d in classes_data.values() if d["overall_avg_pct"] is not None]
        group_avg = round(sum(avgs) / len(avgs), 1) if avgs else None

        ranked = sorted(classes_data.items(), key=lambda x: x[1]["overall_avg_pct"] or 0, reverse=True)
        for rank, (cls_name, _) in enumerate(ranked, 1):
            classes_data[cls_name]["rank_in_group"] = rank
            va = classes_data[cls_name]["overall_avg_pct"]
            classes_data[cls_name]["vs_group_avg"] = (
                round(va - group_avg, 1) if va is not None and group_avg is not None else None
            )

        # ── Topic heatmap: all topics × all classes ───────────────────────
        heatmap = []
        for q in merged:
            row = {
                "topic":         q["topic"],
                "max_marks":     q["max_marks"],
                "class_data":    {},
                "question_text": (question_bank or {}).get(q["topic"], {}).get("text"),
            }
            all_scores = []
            for cls in all_classes:
                scores = q["class_scores"].get(cls, [])
                if scores:
                    p = _pct(scores, q["max_marks"])
                    row["class_data"][cls] = {"avg_pct": p, "num_students": len(scores), "band": _score_band(p)}
                    all_scores.extend(scores)
                else:
                    row["class_data"][cls] = None
            row["group_avg_pct"] = _pct(all_scores, q["max_marks"])
            row["band"]          = _score_band(row["group_avg_pct"])
            heatmap.append(row)
        heatmap.sort(key=lambda x: (x["group_avg_pct"] is None, x["group_avg_pct"] or 0))

        # ── Topic variance ────────────────────────────────────────────────
        topic_variance = []
        for row in heatmap:
            pcts = [d["avg_pct"] for d in row["class_data"].values() if d and d["avg_pct"] is not None]
            if len(pcts) >= 2:
                spread = round(max(pcts) - min(pcts), 1)
                topic_variance.append({"topic": row["topic"], "spread": spread, "group_avg_pct": row["group_avg_pct"]})
        topic_variance.sort(key=lambda x: -x["spread"])

        # ── Topic structure list (shown even when no scores) ─────────────
        # De-duplicate: use heatmap topics if present, else fall back to grid_topics
        if heatmap:
            topic_structure = [
                {"topic": r["topic"], "max_marks": r["max_marks"]}
                for r in heatmap
            ]
        else:
            topic_structure = grid_topics

        total_marks = sum(t["max_marks"] for t in topic_structure)
        total_students = sum(d["student_count"] for d in classes_data.values())

        output["groups"][code] = {
            "name":            group_name,
            "ability":         ability,
            "format":          fmt,
            "has_scores":      has_scores,
            "total_students":  total_students,
            "num_classes":     len(all_classes),
            "group_avg_pct":   group_avg,
            "band":            _score_band(group_avg),
            "all_classes":     all_classes,
            "classes":         classes_data,
            "topic_heatmap":   heatmap,
            "topic_variance":  topic_variance[:5],
            "topic_structure": topic_structure,
            "num_topics":      len(topic_structure),
            "total_marks":     total_marks,
        }

    # ── Cohort summary ────────────────────────────────────────────────────
    output["total_students"]  = sum(g["total_students"] for g in output["groups"].values())
    output["num_groups"]      = len(output["groups"])
    output["num_classes"]     = sum(g["num_classes"]    for g in output["groups"].values())
    output["has_any_scores"]  = any(g["has_scores"] for g in output["groups"].values())
    all_group_avgs = [g["group_avg_pct"] for g in output["groups"].values() if g["group_avg_pct"] is not None]
    output["cohort_avg"]      = round(sum(all_group_avgs) / len(all_group_avgs), 1) if all_group_avgs else None

    # ── Cross-group weak topics (only when scores exist) ─────────────────
    weak_map = {}
    for code, group_data in output["groups"].items():
        for row in group_data["topic_heatmap"]:
            if row["group_avg_pct"] is not None and row["group_avg_pct"] < 55:
                weak_map.setdefault(row["topic"], []).append({
                    "group_code": code, "group_name": group_data["name"], "avg_pct": row["group_avg_pct"],
                })
    output["cross_group_weak"] = sorted(
        [{"topic":         t,
          "groups":        g,
          "min_pct":       min(x["avg_pct"] for x in g),
          "question_text": (question_bank or {}).get(t, {}).get("text")}
         for t, g in weak_map.items() if len(g) >= 2],
        key=lambda x: x["min_pct"],
    )

    return output


def analyse_qla(filepath, group_codes=None, n=5):
    """Lighter analysis used by the generate route — returns weak topics per class."""
    wb       = openpyxl.load_workbook(filepath, data_only=True)
    detected = detect_groups(wb.sheetnames)
    m        = re.search(r"Year\s*(\d+)", str(filepath), re.I)
    year_group = m.group(1) if m else "?"

    selected = group_codes if group_codes else list(detected.keys())
    output   = {"year_group": year_group, "groups": {}}

    for code in selected:
        if code not in detected:
            continue
        group_name, ability, fmt = detected[code]
        sheets      = _sheets_for_group(wb, code)
        all_classes = get_all_classes(wb, sheets)
        group_data  = {"name": group_name, "ability": ability, "format": fmt, "classes": {}}

        for cls in all_classes:
            cf = [cls]
            sheet_qs = [qs for s in sheets for qs in [parse_sheet(wb[s], class_filter=cf)] if qs]
            merged   = merge_questions(sheet_qs)

            weak, all_pcts = [], []
            for q in merged:
                scores = [s for c in cf for s in q["class_scores"].get(c, [])]
                p = _pct(scores, q["max_marks"])
                if p is not None:
                    weak.append({"topic": q["topic"], "max_marks": q["max_marks"], "avg_pct": p, "num_students": len(scores)})
                    all_pcts.append(p)

            weak.sort(key=lambda x: x["avg_pct"])
            overall = round(sum(all_pcts) / len(all_pcts), 1) if all_pcts else None
            group_data["classes"][cls] = {"overall_avg": overall, "weak_topics": weak[:n], "all_topics": weak}

        output["groups"][code] = group_data

    return output
